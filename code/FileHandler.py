''' Team: - Prapat Suriyaphol
        - Masao Yamaguchi
        - Victor Renault

    Research project: Trans-ethnic Genomics Study of Multigenic Disorders
                      by Japan/France International Collaboration

    Project: Set up a database based system to help the study

    Research director: Dr. Fumihiko Matsuda '''

import types
import os
import glob
import gzip

from Utilities import Utilities, getLastErrorMessage

from ObjectBase import *

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Color, Alignment
except ImportError:
    # print 'Warning: module openpyxl could not be loaded'
    Workbook = PatternFill = Color = None


def getFileList(dirName, afterFileName=None):
    fileList = None
    if isinstance(dirName, types.ListType):
        fileList = dirName
    else:
        if ',' in dirName:
            fileList = dirName.split(',')
        elif Utilities.doesDirExist(dirName):
            fileList = glob.glob(os.path.join(dirName, '*'))
        elif Utilities.doesFileExist(dirName):
            fileList = [dirName]
        else:
            fileList = glob.glob(dirName)
    if afterFileName:
        idx = fileList.index(afterFileName)
        print idx, (idx + 1) * 1. / len(fileList)
        fileList = fileList[idx:]
    return fileList


class FileParser(object):
    NO_SPLIT = 0
    _sep = '\t'

    def __init__(self, fileName):
        self.fileName = fileName

    def _getFileHandler(self):
        if isinstance(self.fileName, types.FileType):
            return self.fileName
        # print 'TYPE', type(self.fileName)
        fileExt = Utilities.getFileExtension(self.fileName)
        if fileExt == 'gz':
            # fh = Utilities.popen('cat %s | gunzip' % self.fileName)
            fh = os.popen('gunzip -c %s' % self.fileName)
            # fh = gzip.open(self.fileName)
        elif fileExt == 'xz':
            fh = os.popen('xz -dc ' + self.fileName)
        elif fileExt == 'bz2':
            fh = os.popen('bunzip2 -c ' + self.fileName)
        else:
            fh = open(self.fileName)
        return fh

    def readFileAtOnce(self):
        fh = self._getFileHandler()
        lineList = fh.readlines()
        fh.close()
        return lineList

    def _stripLine(self, line):
        line = '0' + line
        line = line.strip('\n\r')
        return line[1:]

    def getColumnNameToIdxDict(self):
        splittedLine = self.getSplittedLine()
        colDict = dict([[colName, idx]
                        for idx, colName in enumerate(splittedLine)])
        if len(colDict) != len(splittedLine):
            print splittedLine
            raise NotImplementedError
        return colDict

    def getSplittedLine(self, char=None):
        if char is None:
            char = self._sep
        line = self.getNextLine()
        if line is None:
            return
        if char == self.NO_SPLIT:
            return line
        if char is None:
            return line.split()
        return line.split(char)

    def getNextLine(self):
        raise NotImplementedError


class ReadFileAtOnceParser(FileParser):

    def __init__(self, fileName, bufferSize=10000, sep='\t'):
        super(ReadFileAtOnceParser, self).__init__(fileName)
        self.__fh = self._getFileHandler()
        self.__bufferSize = bufferSize
        self.__lineList = self.__getLineListFromFile()
        self._sep = sep

    def __iter__(self):
        while self.hasLinesLeft():
            yield self.getSplittedLine(self._sep)

    def __len__(self):
        return len(self.__lineList)

    def __getLineListFromFile(self):
        if self.__bufferSize:
            return self.__fh.readlines(self.__bufferSize)
        return self.__fh.readlines()

    def getLineList(self):
        if self.__bufferSize:
            raise NotImplementedError
        return self.__lineList

    def getNextLine(self):
        if len(self.__lineList) == 0:
            return
        line = self.popFirst()
        if line is None:
            return
        return self._stripLine(line)

    def hasLinesLeft(self):
        return len(self.__lineList) > 0

    def popFirst(self):
        line = self.__lineList.pop(0)
        if len(self.__lineList) == 0:
            self.__lineList = self.__getLineListFromFile()
        return line

    def restore(self, elt):
        if isinstance(elt, types.ListType):
            elt = self._sep.join(elt)
        self.__lineList = [elt] + self.__lineList


class ExcelCell:

    def __init__(self, value, color=None, bgColor=None, style=None,
                 alignment=None, otherParamDict=None):
        self.value = value
        self.color = color
        self.bgColor = bgColor
        self.style = style
        self.alignment = alignment
        self.otherParamDict = otherParamDict


class CsvFileWriter:

    def __init__(self, fileName, fileOption='w', exportInExcel=False,
                 useBuffer=True, otherParamDict=None):
        self._fileName = fileName
        self.__isExcel = False
        self.__otherParamDict = otherParamDict
        if isinstance(fileName, types.FileType):
            self.__fh = fileName
            exportInExcel = False
        else:
            fileExt = Utilities.getFileExtension(fileName)
            if fileExt == 'gz':
                # self.__fh = gzip.open(fileName, fileOption)
                # from subprocess import Popen,PIPE
                # GZ = Popen('gzip > %s' % fileName,stdin=PIPE,shell=True)
                # self.__fh = GZ.stdin
                # self.__gz = GZ
                self.__fh = open(fileName[:-3], fileOption)
            elif fileExt == 'xlsx':
                self.__wb = Workbook(guess_types=True)
                self.__fh = self.__wb.active
                if otherParamDict and 'title' in otherParamDict:
                    self.__fh.title = otherParamDict['title']
                self.__isExcel = True
                self.__lineNb = 1
            else:
                self.__fh = open(fileName, fileOption)
            if Utilities.getFileExtension(fileName) == 'xls':
                exportInExcel = False
        self.__exportInExcel = exportInExcel
        self.__useBuffer = useBuffer
        self.__buffer = []

    def _addNewWorkSheet(self, title=None):
        paramDict = {}
        if title:
            paramDict['title'] = title
        self.__fh = self.__wb.create_sheet(**paramDict)
        self.__lineNb = 1

    def __del__(self):
        if self.__isExcel:
            self.__wb.save(self._fileName)
            return
        if not hasattr(self.__fh, 'closed') or not self.__fh.closed:
            self.close()
        if self.__exportInExcel:
            self.exportInExcelFormat()

    def writeListsInColumn(self, *args):
        lineToWriteList = []
        while True:
            lineToWrite = []
            isFinished = True
            for list in args:
                if len(list):
                    value = list.pop(0)
                    isFinished = False
                else:
                    value = ''
                lineToWrite.append(value)
            if isFinished:
                break
            lineToWriteList.append(lineToWrite)
        self.writeAllLinesAtOnce(lineToWriteList)

    def __convertAllEltInListToString(self, eltList):
        if isinstance(eltList, types.StringType):
            return [eltList]
        return [str(elt) for elt in eltList]

    def _writeAndEmptyBuffer(self):
        self.__fh.write('\n'.join(self.__buffer) + '\n')
        self.__buffer = []

    def __writeExcelLine(self, fieldList):
        line = []
        cellToChangeDict = {}
        for i, field in enumerate(fieldList):
            if i < 26:
                column = chr(ord('A') + i)
            else:
                column = 'A' + chr(ord('A') + i % 26)
            if isinstance(field, ExcelCell):
                cellToChangeDict['%s%d' % (
                    column, self.__lineNb)] = field.color, field.bgColor,
                field.style, field.alignment, field.otherParamDict
                field = field.value
            line.append(field)
        self.__fh.append(line)
        for cellKey, (color, bgColor, style, alignment, otherParamDict) in \
                cellToChangeDict.iteritems():
            if color or style:
                paramDict = {}
                if color:
                    paramDict['color'] = color
                if style:
                    if style == 'italic':
                        paramDict['italic'] = True
                    elif style == 'bold':
                        paramDict['bold'] = True
                if otherParamDict:
                    paramDict.update(otherParamDict)
                try:
                    newFont = self.__fh[cellKey].font.copy(**paramDict)
                except:
                    print paramDict, cellKey
                    print len(fieldList), fieldList
                    raise
                self.__fh[cellKey].font = newFont
            if bgColor:
                self.__fh[cellKey].fill = PatternFill(
                    "solid", fgColor=Color(bgColor))
            if alignment or (self.__otherParamDict and
                             'alignment' in self.__otherParamDict):
                if not alignment:
                    alignment = self.__otherParamDict['alignment']
                self.__fh[cellKey].alignment = Alignment(**alignment)
        self.__lineNb += 1

    def write(self, fieldList, sep='\t'):
        if self.__isExcel:
            self.__writeExcelLine(fieldList)
            return
        fieldList = self.__convertAllEltInListToString(fieldList)
        fieldStr = sep.join(fieldList)
        if self.__useBuffer:
            self.__buffer.append(fieldStr)
        else:
            self.__fh.write(fieldStr + '\n')
        if self.__useBuffer and len(self.__buffer) >= 20000:
            self._writeAndEmptyBuffer()

    def writeAllLinesAtOnce(self, lineList, sep='\t'):
        lineList = [sep.join(self.__convertAllEltInListToString(
            fieldList)) for fieldList in lineList]
        self.__fh.write('\n'.join(lineList) + '\n')

    def close(self):
        if self.__isExcel:
            self.__wb.save(self._fileName)
            return
        if len(self.__buffer):
            self._writeAndEmptyBuffer()
        if not hasattr(self.__fh, 'closed') or not self.__fh.closed:
            self.__fh.close()
        try:
            self.__fh.closed = True
        except AttributeError:
            pass
        if self._fileName[-3:] == '.gz':
            # self.__gz.wait()
            Utilities.mySystem('gzip -f %s' % self._fileName[:-3])

    def exportInExcelFormat(self):
        if not self.__fh.closed:
            self.close()
        self.excelfileName = Utilities.createExcelFileFromFileAndReturnName(
            self._fileName)


class SaveFileInRepository:

    def getArchiveFileName(self, obj):
        dateTag = obj.getDateTag()
        timeStampStr = dateTag.strftime("%Y%m%d_%H%M%S")
        return '.'.join([obj.getFileName(), timeStampStr])

    def save(self, obj):
        obj._archiveFileName = self.getArchiveFileName(obj)
        newFileName = os.path.join(obj.getSaveDir(), obj._archiveFileName)
        try:
            os.rename(obj.getOriginalFileName(), newFileName)
        except:
            print 'Can not rename file from "%s" to "%s"' % (
                obj.getOriginalFileName(), newFileName)
            raise


class FileNameGetter:

    def __init__(self, fileName):
        self.__fileName = fileName
        self.__fileExt = Utilities.getFileExtension(fileName)

    def get(self, newFileExt, checkFileNames = True):
        shift = 0
        if newFileExt[0] in ['_', '.']:
            shift = -1
        newFileName = self.__fileName[
            :-len(self.__fileExt) + shift] + newFileExt
        if checkFileNames and newFileName == self.__fileName:
            raise NotImplementedError('new file name is the same as original \
one: file = "%s", fileExt = [%s], newFileExt = [%s]' % (self.__fileName,
                                                        self.__fileExt,
                                                        newFileExt))
        #if '..' in newFileName:
            #print self.__fileName, self.__fileExt, newFileExt, newFileName
            #raise NotImplementedError
        return newFileName


class FastaSeq:

    def __init__(self, seqName, seq):
        self.seqName = seqName
        self.seq = seq

    def getShortSeqName(self):
        partList = self.seqName.split('|')
        seqName = self.seqName
        if partList[0] == '>gi':
            seqName = '|'.join([partList[1], partList[3]])
        else:
            seqName = seqName.lstrip('>').strip().split()[0]
        return seqName.lstrip('>').strip()

    def __len__(self):
        return len(self.seq)

    def writeInFastQ(self, step, fh, fh2=None, insertSize=None, test=None):
        qualityStr = 'I' * step
        seqName = self.getShortSeqName()
        end = len(self) - step + 1
        if insertSize:
            end -= (insertSize + step)
        for i in xrange(end):
            header = '%s|%d-%d' % (seqName, i + 1, i + step)
            header += '/%d'
            for line in ['@%s' % (header % 1), self.seq[i:i + step], '+',
                         qualityStr]:
                fh.write(line)
            if fh2:
                seq2 = self.seq[i + step +
                                insertSize:i + 2 * step + insertSize]
                if not test:
                    seq2 = Sequence.getReverseComplementString(seq2)
                for line in ['@%s' % (header % (2)), seq2, '+', qualityStr]:
                    fh2.write(line)


class ParseFastaFile(object):
    '''
    classdocs
    '''

    def __init__(self, name):
        '''
        Constructor
        '''
        self.filename = name
        self.sequences = {}

    def _getFastaSeqForChr(self, chrName):
        fh = ReadFileAtOnceParser(self.filename)
        while fh.hasLinesLeft():
            fastaSeq = self._getNextSeqFromFh(fh)
            if not fastaSeq:
                break
            fastaSeqChrName = fastaSeq.getShortSeqName().split('_')[0]
            if fastaSeqChrName != chrName:
                print 'Passing [%s]' % fastaSeqChrName
                continue
            return fastaSeq

    def _getNextSeqFromFh(self, fh, allowEmptyLines=False):
        header = fh.popFirst().strip()
        if not header:
            return
        seq = ''
        while fh.hasLinesLeft():
            currentSeq = fh.popFirst().strip()
            if allowEmptyLines and not currentSeq:
                break
            if currentSeq[0] == '>':
                fh.restore(currentSeq)
                break
            seq += currentSeq
        if header[0] != '>':
            raise NotImplementedError(
                'Expecting ">" as 1st header character but found [%s]' %
                header)
        return FastaSeq(header, seq)

    def _getIdxFileName(self):
        return self.filename + '_idx'

    def _createFastaIdxFile(self):
        idxFile = self._getIdxFileName()
        if not os.path.isfile(idxFile):
            Utilities.mySystem('grep ">" %s > %s' % (self.filename, idxFile))

    def _getSequenceNameListFromFile(self, filtered=False, excludeMT=False):
        self._createFastaIdxFile()
        fh = open(self._getIdxFileName())
        for seqName in fh:
            seqName = seqName.lstrip('>').strip().split()[0]
            if filtered:
                seqName = FastaSeq(
                    seqName, None).getShortSeqName().split('_')[0]
                hasChrPrefix = seqName[:3] == 'chr'
                if (not hasChrPrefix and
                    (not ValueParser().isNb(seqName) and len(seqName) not in
                     [1, 2])) or (hasChrPrefix and
                                  not ValueParser().isNb(seqName[3:]) and
                                  len(seqName) not in [4, 5]):
                    continue
                if not ValueParser().isNb(seqName.replace('chr', '')) and seqName.replace('chr', '') not in ['X', 'Y', 'M', 'MT']:
                    continue
            if excludeMT and seqName in ['M', 'MT', 'chrM', 'chrMT']:
                continue
            yield seqName
        fh.close()

    def parse(self):
        descriptions = {}
        if self.filename.split('.')[-1] == 'gz':
            import gzip
            infile = gzip.open(self.filename, 'r')
        else:
            infile = open(self.filename, 'r')
        lines = infile.readlines()
        infile.close()
        curDescription = ''
        for l in lines:
            if l.startswith('#') or l == '\n':
                pass
            elif l.startswith('>'):
                print l
                curDescription = l.replace('>', '')
                self.sequences[curDescription] = ''
                name = l.split('|')[3]
                if name not in descriptions:
                    descriptions[name] = 1
                else:
                    descriptions[name] += 1
            else:
                self.sequences[curDescription] += l.replace('\n', '')
        for d in self.sequences:
            print d


class MergeAnnotationFiles:

    def __getChrListFromIdxFile(self, fileName):
        fh = ReadFileAtOnceParser(fileName)
        chrList = []
        for splittedLine in fh:
            chrList.append(splittedLine[0])
        return chrList

    def __getChrListFromRefFile(self, fileName):
        if Utilities.getFileExtension(fileName) == 'sizes':
            return self.__getChrListFromIdxFile(fileName)
        idxFileName = fileName + '_idx'
        if not os.path.isfile(idxFileName):
            cmd = 'grep ">" %s > %s' % (fileName, idxFileName)
            Utilities.mySystem(cmd)
        fh = ReadFileAtOnceParser(idxFileName)
        chrList = []
        for splittedLine in fh:
            chrName = splittedLine[0].lstrip('>').split('|')[0].strip()
            chrList.append(chrName)
        return chrList

    def __isRefFileForHuman(self, fileName):
        partList = fileName.split(os.path.sep)
        for part in partList:
            part = part.lower()
            if 'homo' in part and 'sapiens' in part:
                return True

    def __getFileListFromPatternAndChrList(self, pattern, chrList):
        fileList = []
        for chrName in chrList:
            fileName = pattern % chrName
            if os.path.isfile(fileName):
                fileList.append(fileName)
        return fileList

    def process(self, pattern, outputFileName, refFileName=None,
                hasHeader=None, chrNameList=None):
        chrList = range(1, 23) + ['X', 'Y', 'MT', 'M']
        if refFileName and not self.__isRefFileForHuman(refFileName):
            chrList = self.__getChrListFromRefFile(refFileName)
        if chrNameList:
            chrList = chrNameList
        fileList = self.__getFileListFromPatternAndChrList(pattern, chrList)
        if not fileList and chrList and chrList[0][:3] == 'chr':
            fileList = self.__getFileListFromPatternAndChrList(
                pattern, [chrName[3:] for chrName in chrList])
        # os.system('cat %s > %s' % (' '.join(fileList), outputFileName))
        toCompress = False
        if outputFileName[-3:] == '.gz':
            outputFileName = '.'.join(outputFileName.split('.')[:-1])
            toCompress = True
        if hasHeader:
            cmd = 'head -1 %s > %s' % (fileList[0], outputFileName)
            if os.path.basename(fileList[0]).split('.')[-1] == 'gz':
                cmd = 'zcat %s | head -1 > %s' % (fileList[0], outputFileName)
            Utilities.mySystem(cmd)
        for fileName in fileList:
            if os.path.basename(fileName).split('.')[-1] == 'gz':
                cmd = 'zcat'
                if hasHeader:
                    cmd = "zcat %s | sed -n '1!p' >> %s" % (
                        fileName, outputFileName)
                else:
                    cmd += ' %s >> %s' % (fileName, outputFileName)
            else:
                cmd = 'cat'
                if hasHeader:
                    cmd = "sed -n '1!p'"
                cmd += ' %s >> %s' % (fileName, outputFileName)
            Utilities.mySystem(cmd)
        if toCompress:
            Utilities.mySystem('gzip %s' % outputFileName)


def run(options, args):
    print FileNameGetter(options.fileName).get(options.fileExt)

runFromTerminal(__name__, [CommandParameter('f', 'fileName', 'string'),
                           CommandParameter('F', 'fileExt', 'string')], run)
