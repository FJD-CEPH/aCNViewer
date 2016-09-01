''' Team: - Prapat Suriyaphol
		- Masao Yamaguchi
		- Victor Renault
				
	Research project: Trans-ethnic Genomics Study of Multigenic Disorders
										by Japan/France International Collaboration
						
	Project: Set up a database based system to help the study
				
	Research director: Dr. Fumihiko Matsuda '''

import types
import os
import glob
import gzip

from Utilities import Utilities, getLastErrorMessage

try:
	from openpyxl import Workbook
	from openpyxl.styles import PatternFill, Color, Alignment
except ImportError:
	print 'Warning: module openpyxl could not be loaded'
	Workbook = PatternFill = Color = None


def getFileList(dirName, afterFileName = None):
	fileList = None
	if type(dirName) == types.ListType:
		fileList = dirName
	else:
		if ',' in dirName:
			fileList = dirName.split(',')
		elif Utilities.doesDirExist(dirName):
			fileList = glob.glob(os.path.join(dirName, '*'))
		elif Utilities.doesFileExist(dirName):
			fileList = [dirName]
		else:
			fileList = glob.glob(dirName)
	if afterFileName:
		idx = fileList.index(afterFileName)
		print idx, (idx + 1) * 1. / len(fileList)
		fileList = fileList[idx:]
	return fileList


class FileParser(object):
	NO_SPLIT = 0
	_sep = '\t'
	
	def __init__(self, fileName):
		self.fileName = fileName
	
	def _getFileHandler(self):
		if type(self.fileName) == types.FileType:
			return self.fileName
		#print 'TYPE', type(self.fileName)
		fileExt = Utilities.getFileExtension(self.fileName)
		if fileExt == 'gz':
			#fh = Utilities.popen('cat %s | gunzip' % self.fileName)
			fh = os.popen('gunzip -c %s' % self.fileName)
			#fh = gzip.open(self.fileName)
		elif fileExt == 'xz':
			fh = os.popen('xz -dc ' + self.fileName)
		elif fileExt == 'bz2':
			fh = os.popen('bunzip2 -c ' + self.fileName)
		else:
			fh = open(self.fileName)
		return fh
		
	def readFileAtOnce(self):
		fh = self._getFileHandler()
		lineList = fh.readlines()
		fh.close()
		return lineList
	
	def _stripLine(self, line):
		line = '0' + line
		line = line.strip('\n\r')
		return line[1:]
	
	def getColumnNameToIdxDict(self):
		splittedLine = self.getSplittedLine()
		colDict = dict([[colName, idx] for idx, colName in enumerate(splittedLine)])
		if len(colDict) != len(splittedLine):
			print splittedLine
			raise NotImplementedError
		return colDict
	
	def getSplittedLine(self, char = None):
		if char is None:
			char = self._sep
		line = self.getNextLine()
		if line is None:
			return
		if char == self.NO_SPLIT:
			return line
		if char is None:
			return line.split()
		return line.split(char)
	
	def getNextLine(self):
		raise NotImplementedError


class ReadFileAtOnceParser(FileParser):
	def __init__(self, fileName, bufferSize = 10000, sep = '\t'):
		super(ReadFileAtOnceParser, self).__init__(fileName)
		self.__fh = self._getFileHandler()
		self.__bufferSize = bufferSize
		self.__lineList = self.__getLineListFromFile()
		self._sep = sep
	
	def __iter__(self):
		while self.hasLinesLeft():
			yield self.getSplittedLine(self._sep)
			
	def __len__(self):
		return len(self.__lineList)
			
	def __getLineListFromFile(self):
		if self.__bufferSize:
			return self.__fh.readlines(self.__bufferSize)
		return self.__fh.readlines()
	
	def getLineList(self):
		if self.__bufferSize:
			raise NotImplementedError
		return self.__lineList
	
	def getNextLine(self):
		if len(self.__lineList) == 0:
			return
		line = self.popFirst()
		if line is None:
			return
		return self._stripLine(line)
	
	def hasLinesLeft(self):
		return len(self.__lineList) > 0
	
	def popFirst(self):
		line = self.__lineList.pop(0)
		if len(self.__lineList) == 0:
			self.__lineList = self.__getLineListFromFile()
		return line
	
	def restore(self, elt):
		if type(elt) == types.ListType:
			elt = self._sep.join(elt)
		self.__lineList = [elt] + self.__lineList

class ExcelCell:
	def __init__(self, value, color = None, bgColor = None, style = None, alignment = None, otherParamDict = None):
		self.value = value
		self.color = color
		self.bgColor = bgColor
		self.style = style
		self.alignment = alignment
		self.otherParamDict = otherParamDict


class CsvFileWriter:
	def __init__(self, fileName, fileOption = 'w', exportInExcel = False, useBuffer = True, otherParamDict = None):
		self._fileName = fileName
		self.__isExcel = False
		self.__otherParamDict = otherParamDict
		if type(fileName) == types.FileType:
			self.__fh = fileName
			exportInExcel = False
		else:
			fileExt = Utilities.getFileExtension(fileName)
			if fileExt == 'gz':
				#self.__fh = gzip.open(fileName, fileOption)
				#from subprocess import Popen,PIPE
				#GZ = Popen('gzip > %s' % fileName,stdin=PIPE,shell=True)
				#self.__fh = GZ.stdin
				#self.__gz = GZ
				self.__fh = open(fileName[:-3], fileOption)
			elif fileExt == 'xlsx':
				self.__wb = Workbook(guess_types = True)
				self.__fh = self.__wb.active
				if otherParamDict and otherParamDict.has_key('title'):
					self.__fh.title = otherParamDict['title']
				self.__isExcel = True
				self.__lineNb = 1
			else:
				self.__fh = open(fileName, fileOption)
			if Utilities.getFileExtension(fileName) == 'xls':
				exportInExcel = False
		self.__exportInExcel = exportInExcel
		self.__useBuffer = useBuffer
		self.__buffer = []
	
	def _addNewWorkSheet(self, title = None):
		paramDict = {}
		if title:
			paramDict['title'] = title
		self.__fh = self.__wb.create_sheet(**paramDict)
		self.__lineNb = 1

	def __del__(self):
		if self.__isExcel:
			self.__wb.save(self._fileName)
			return
		if not hasattr(self.__fh, 'closed') or not self.__fh.closed:
			self.close()
		if self.__exportInExcel:
			self.exportInExcelFormat()
	
	def writeListsInColumn(self, *args):
		lineToWriteList = []
		while True:
			lineToWrite = []
			isFinished = True
			for list in args:
				if len(list):
					value = list.pop(0)
					isFinished = False
				else:
					value = ''
				lineToWrite.append(value)
			if isFinished:
				break
			lineToWriteList.append(lineToWrite)
		self.writeAllLinesAtOnce(lineToWriteList)
	
	def __convertAllEltInListToString(self, eltList):
		if type(eltList) == types.StringType:
			return [eltList]
		return [str(elt) for elt in eltList]
	
	def _writeAndEmptyBuffer(self):
		self.__fh.write('\n'.join(self.__buffer) + '\n')
		self.__buffer = []
	
	def __writeExcelLine(self, fieldList):
		line = []
		cellToChangeDict = {}
		for i, field in enumerate(fieldList):
			if isinstance(field, ExcelCell):
				cellToChangeDict['%s%d' % (chr(ord('A')+i), self.__lineNb)] = field.color, field.bgColor, field.style, field.alignment, field.otherParamDict
				field = field.value
			line.append(field)
		self.__fh.append(line)
		for cellKey, (color, bgColor, style, alignment, otherParamDict) in cellToChangeDict.iteritems():
			if color or style:
				paramDict = {}
				if color:
					paramDict['color'] = color
				if style:
					if style == 'italic':
						paramDict['italic'] = True
					elif style == 'bold':
						paramDict['bold'] = True
				if otherParamDict:
					paramDict.update(otherParamDict)
				newFont = self.__fh[cellKey].font.copy(**paramDict)
				self.__fh[cellKey].font = newFont
			if bgColor:
				self.__fh[cellKey].fill = PatternFill("solid", fgColor=Color(bgColor))
			if alignment or (self.__otherParamDict and self.__otherParamDict.has_key('alignment')):
				if not alignment:
					alignment = self.__otherParamDict['alignment']
				self.__fh[cellKey].alignment = Alignment(**alignment)
		self.__lineNb += 1

	def write(self, fieldList, sep = '\t'):
		if self.__isExcel:
			self.__writeExcelLine(fieldList)
			return
		fieldList = self.__convertAllEltInListToString(fieldList)
		fieldStr = sep.join(fieldList)
		if self.__useBuffer:
			self.__buffer.append(fieldStr)
		else:
			self.__fh.write(fieldStr + '\n')
		if self.__useBuffer and len(self.__buffer) >= 20000:
			self._writeAndEmptyBuffer()
	
	def writeAllLinesAtOnce(self, lineList, sep = '\t'):
		lineList = [sep.join(self.__convertAllEltInListToString(fieldList)) for fieldList in lineList]
		self.__fh.write('\n'.join(lineList) + '\n')
		
	def close(self):
		if self.__isExcel:
			self.__wb.save(self._fileName)
			return
		if len(self.__buffer):
			self._writeAndEmptyBuffer()
		if not hasattr(self.__fh, 'closed') or not self.__fh.closed:
			self.__fh.close()
		try:
			self.__fh.closed = True
		except AttributeError:
			pass
		if self._fileName[-3:] == '.gz':
			#self.__gz.wait()
			Utilities.mySystem('gzip -f %s' % self._fileName[:-3])
	
	def exportInExcelFormat(self):
		if not self.__fh.closed:
			self.close()
		self.excelfileName = Utilities.createExcelFileFromFileAndReturnName(self._fileName)


class SaveFileInRepository:
	def getArchiveFileName(self, obj):
		dateTag = obj.getDateTag()
		timeStampStr = dateTag.strftime("%Y%m%d_%H%M%S")
		return '.'.join([obj.getFileName(), timeStampStr])
	
	def save(self, obj):
		obj._archiveFileName = self.getArchiveFileName(obj)
		newFileName = os.path.join(obj.getSaveDir(), obj._archiveFileName)
		try:
			os.rename(obj.getOriginalFileName(), newFileName)
		except:
			print 'Can not rename file from "%s" to "%s"' % (obj.getOriginalFileName(), newFileName)
			raise
		
		
class FileNameGetter:
	def __init__(self, fileName):
		self.__fileName = fileName
		self.__fileExt = Utilities.getFileExtension(fileName)
	
	def get(self, newFileExt):
		shift = 0
		if newFileExt[0] in ['_', '.']:
			shift = -1
		newFileName = self.__fileName[:-len(self.__fileExt)+shift] + newFileExt
		if newFileName == self.__fileName:
			raise NotImplementedError('new file name is the same as original one: file = "%s", fileExt = [%s], newFileExt = [%s]' % (self.__fileName, self.__fileExt, newFileExt))
		if '..' in newFileName:
			print self.__fileName, self.__fileExt, newFileExt, newFileName
			raise NotImplementedError
		return newFileName
		
	
from ObjectBase import *
	
def run(options, args):
	print FileNameGetter(options.fileName).get(options.fileExt)

runFromTerminal(__name__, [CommandParameter('f', 'fileName', 'string'),
                           CommandParameter('F', 'fileExt', 'string')], run)
